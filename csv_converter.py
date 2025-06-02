import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import zipfile
import tempfile
import shutil

def csv_to_formatted_xlsx(csv_file_path, xlsx_file_path=None, table_style="TableStyleMedium9"):
    """
    CSV dosyasını formatlanmış XLSX'e dönüştürür
    """
    if xlsx_file_path is None:
        base_name = os.path.splitext(csv_file_path)[0]
        xlsx_file_path = f"{base_name}_formatted.xlsx"
    
    try:
        # CSV dosyasını oku
        df = pd.read_csv(csv_file_path, encoding='utf-8')
        
        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # DataFrame'i Excel'e yaz
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Tablo oluştur
        table_range = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df) + 1}"
        table = Table(displayName="DataTable", ref=table_range)
        
        # Tablo stilini ayarla
        style = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Sütun genişliklerini içeriğe tam uyacak şekilde ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value is not None:
                        # String uzunluğunu hesapla
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Sütun genişliğini içeriğin tam sığacağı kadar ayarla
            if max_length > 0:
                # Her karakter için yaklaşık 1.1 birim + 2 birim padding
                adjusted_width = (max_length * 1.1) + 2
                # Minimum genişlik 8, maksimum sınır yok (çok uzun içerikler için)
                adjusted_width = max(adjusted_width, 8)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Tüm hücreleri sola yasla (text wrapping KAPALI)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal='left', 
                    vertical='center',
                    wrap_text=False  # Text wrapping kapalı - tek satırda göster
                )
        
        # Excel dosyasını kaydet
        wb.save(xlsx_file_path)
        return True, xlsx_file_path, len(df), len(df.columns)
        
    except Exception as e:
        return False, str(e), 0, 0

def extract_zip_and_get_csvs(zip_path):
    """
    ZIP dosyasını çıkartır ve CSV dosyalarını bulur
    """
    try:
        # Geçici klasör oluştur
        temp_dir = tempfile.mkdtemp(prefix="csv_converter_")
        
        # ZIP dosyasını çıkart
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # CSV dosyalarını bul (tüm alt klasörlerde)
        csv_files = []
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.lower().endswith('.csv'):
                    full_path = os.path.join(root, file)
                    csv_files.append((full_path, file))
        
        return temp_dir, csv_files
        
    except Exception as e:
        return None, str(e)

class CSVConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV to Excel Converter (ZIP Destekli)")
        self.root.geometry("650x500")
        self.root.configure(bg="#f0f0f0")
        
        # Ana frame
        main_frame = tk.Frame(root, bg="#f0f0f0", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Başlık
        title_label = tk.Label(
            main_frame, 
            text="📊 CSV to Excel Converter", 
            font=("Arial", 16, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50"
        )
        title_label.pack(pady=(0, 10))
        
        # Alt başlık
        subtitle_label = tk.Label(
            main_frame,
            text="ZIP dosyaları ve CSV'leri Excel tablosuna dönüştürün",
            font=("Arial", 10),
            bg="#f0f0f0",
            fg="#7f8c8d"
        )
        subtitle_label.pack(pady=(0, 25))
        
        # Buton frame
        button_frame = tk.Frame(main_frame, bg="#f0f0f0")
        button_frame.pack(pady=10)
        
        # ZIP dosyası seçme butonu
        self.zip_btn = tk.Button(
            button_frame,
            text="📦 ZIP Dosyası Seç",
            command=self.select_zip_file,
            font=("Arial", 11, "bold"),
            bg="#e74c3c",
            fg="white",
            relief="flat",
            padx=25,
            pady=8,
            cursor="hand2"
        )
        self.zip_btn.pack(side="left", padx=5)
        
        # Tek CSV dosyası seçme butonu
        self.select_btn = tk.Button(
            button_frame,
            text="📁 CSV Dosyası Seç",
            command=self.select_single_file,
            font=("Arial", 11, "bold"),
            bg="#3498db",
            fg="white",
            relief="flat",
            padx=25,
            pady=8,
            cursor="hand2"
        )
        self.select_btn.pack(side="left", padx=5)
        
        # Klasör seçme butonu
        self.batch_btn = tk.Button(
            button_frame,
            text="📂 Klasör Seç",
            command=self.select_folder,
            font=("Arial", 11, "bold"),
            bg="#27ae60",
            fg="white",
            relief="flat",
            padx=25,
            pady=8,
            cursor="hand2"
        )
        self.batch_btn.pack(side="left", padx=5)
        
        # Renk seçimi frame
        color_frame = tk.Frame(main_frame, bg="#f0f0f0")
        color_frame.pack(pady=15)
        
        # Renk seçimi etiketi
        color_label = tk.Label(
            color_frame,
            text="🎨 Tablo Rengi:",
            font=("Arial", 10, "bold"),
            bg="#f0f0f0",
            fg="#2c3e50"
        )
        color_label.pack(side="left", padx=(0, 10))
        
        # Renk seçenekleri
        self.table_colors = {
            "🔵 Mavi (Varsayılan)": "TableStyleMedium9",
            "🟠 Turuncu": "TableStyleMedium2", 
            "🟢 Yeşil": "TableStyleMedium7",
            "🟣 Mor": "TableStyleMedium15"
        }
        
        self.color_var = tk.StringVar(value="🔵 Mavi (Varsayılan)")
        self.color_dropdown = ttk.Combobox(
            color_frame,
            textvariable=self.color_var,
            values=list(self.table_colors.keys()),
            state="readonly",
            width=20,
            font=("Arial", 9)
        )
        self.color_dropdown.pack(side="left")
        
        # Progress bar
        self.progress = ttk.Progressbar(
            main_frame,
            mode='indeterminate',
            length=350
        )
        self.progress.pack(pady=20)
        self.progress.pack_forget()  # Başlangıçta gizle
        
        # Durum etiketi
        self.status_label = tk.Label(
            main_frame,
            text="Dosya veya ZIP seçin...",
            font=("Arial", 10),
            bg="#f0f0f0",
            fg="#34495e"
        )
        self.status_label.pack(pady=10)
        
        # Sonuç alanı
        self.result_frame = tk.Frame(main_frame, bg="#f0f0f0")
        self.result_frame.pack(fill="both", expand=True, pady=20)
        
        self.result_text = tk.Text(
            self.result_frame,
            height=10,
            font=("Consolas", 9),
            bg="#2c3e50",
            fg="#ecf0f1",
            relief="flat",
            padx=10,
            pady=10
        )
        self.result_text.pack(fill="both", expand=True)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(self.result_frame)
        scrollbar.pack(side="right", fill="y")
        self.result_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.result_text.yview)
        
        self.log("🚀 Program hazır!")
        self.log("📦 ZIP dosyası, 📁 tek CSV veya 📂 klasör seçebilirsiniz.")
        self.log("🎨 Tablo rengini değiştirmeyi unutmayın!")
    
    def log(self, message):
        """Sonuç alanına mesaj ekle"""
        self.result_text.insert(tk.END, f"{message}\n")
        self.result_text.see(tk.END)
        self.root.update()
    
    def show_progress(self):
        """Progress bar göster ve başlat"""
        self.progress.pack(pady=20)
        self.progress.start()
    
    def hide_progress(self):
        """Progress bar gizle ve durdur"""
        self.progress.stop()
        self.progress.pack_forget()
    
    def get_selected_table_style(self):
        """Seçilen tablo stilini döndür"""
        selected_color = self.color_var.get()
        return self.table_colors.get(selected_color, "TableStyleMedium9")
    
    def select_zip_file(self):
        """ZIP dosyası seçimi"""
        file_path = filedialog.askopenfilename(
            title="ZIP Dosyası Seçin",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        
        if file_path:
            self.status_label.config(text="ZIP çıkartılıyor ve dönüştürülüyor...")
            self.show_progress()
            
            # Threading kullanarak UI donmasını engelle
            thread = threading.Thread(target=self.process_zip_file, args=(file_path,))
            thread.start()
    
    def process_zip_file(self, zip_path):
        """ZIP dosyası işleme"""
        temp_dir = None
        try:
            self.log(f"📦 ZIP dosyası işleniyor: {os.path.basename(zip_path)}")
            
            # ZIP'i çıkart ve CSV'leri bul
            temp_dir, csv_files = extract_zip_and_get_csvs(zip_path)
            
            if temp_dir is None:
                self.log(f"❌ ZIP çıkartma hatası: {csv_files}")
                self.status_label.config(text="❌ ZIP hatası!")
                messagebox.showerror("Hata", f"ZIP dosyası çıkartılamadı:\n{csv_files}")
                return
            
            if not csv_files:
                self.log("❌ ZIP içinde CSV dosyası bulunamadı!")
                self.status_label.config(text="❌ CSV bulunamadı!")
                messagebox.showwarning("Uyarı", "ZIP dosyası içinde CSV dosyası bulunamadı!")
                return
            
            self.log(f"✅ {len(csv_files)} CSV dosyası bulundu")
            
            # ZIP ile aynı klasöre kaydet (klasör oluşturmadan)
            zip_dir = os.path.dirname(zip_path)
            table_style = self.get_selected_table_style()
            
            # CSV'leri dönüştür
            success_count = 0
            total_rows = 0
            
            for i, (csv_path, csv_name) in enumerate(csv_files):
                self.log(f"\n🔄 ({i+1}/{len(csv_files)}) {csv_name}")
                
                # Çıktı dosya yolu (direkt ZIP'in bulunduğu klasöre)
                xlsx_name = os.path.splitext(csv_name)[0] + "_formatted.xlsx"
                xlsx_path = os.path.join(zip_dir, xlsx_name)
                
                success, result, rows, cols = csv_to_formatted_xlsx(csv_path, xlsx_path, table_style)
                
                if success:
                    self.log(f"✅ Başarılı: {rows} satır, {cols} sütun")
                    success_count += 1
                    total_rows += rows
                else:
                    self.log(f"❌ Hata: {result}")
            
            self.log(f"\n🎉 ZIP işlemi tamamlandı!")
            self.log(f"✅ Başarılı: {success_count}/{len(csv_files)} dosya")
            self.log(f"📊 Toplam işlenen satır: {total_rows}")
            self.log(f"📁 XLSX dosyaları ZIP ile aynı klasöre kaydedildi")
            
        except Exception as e:
            self.log(f"❌ ZIP işlemi hatası: {str(e)}")
            self.status_label.config(text="❌ Hata oluştu!")
            messagebox.showerror("Hata", f"ZIP işlemi hatası:\n{str(e)}")
        
        finally:
            # Geçici klasörü temizle
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                except:
                    pass
            
            self.hide_progress()
    
    def select_single_file(self):
        """Tek dosya seçimi"""
        file_path = filedialog.askopenfilename(
            title="CSV Dosyası Seçin",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if file_path:
            self.status_label.config(text="Dönüştürülüyor...")
            self.show_progress()
            
            # Threading kullanarak UI donmasını engelle
            thread = threading.Thread(target=self.convert_single_file, args=(file_path,))
            thread.start()
    
    def convert_single_file(self, file_path):
        """Tek dosya dönüştürme işlemi"""
        try:
            self.log(f"📄 Dönüştürülüyor: {os.path.basename(file_path)}")
            
            table_style = self.get_selected_table_style()
            success, result, rows, cols = csv_to_formatted_xlsx(file_path, None, table_style)
            
            if success:
                self.log(f"✅ Başarılı! {rows} satır, {cols} sütun işlendi")
                self.log(f"💾 Kaydedildi: {os.path.basename(result)}")
                self.status_label.config(text="✅ Dönüştürme tamamlandı!")
                messagebox.showinfo("Başarılı", f"Dosya başarıyla dönüştürüldü!\n\n{os.path.basename(result)}")
            else:
                self.log(f"❌ Hata: {result}")
                self.status_label.config(text="❌ Hata oluştu!")
                messagebox.showerror("Hata", f"Dönüştürme sırasında hata:\n{result}")
                
        except Exception as e:
            self.log(f"❌ Beklenmeyen hata: {str(e)}")
            self.status_label.config(text="❌ Hata oluştu!")
            messagebox.showerror("Hata", f"Beklenmeyen hata:\n{str(e)}")
        
        finally:
            self.hide_progress()
    
    def select_folder(self):
        """Klasör seçimi (toplu işlem)"""
        folder_path = filedialog.askdirectory(title="CSV Dosyalarının Bulunduğu Klasörü Seçin")
        
        if folder_path:
            csv_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.csv')]
            
            if not csv_files:
                messagebox.showwarning("Uyarı", "Seçilen klasörde CSV dosyası bulunamadı!")
                return
            
            self.status_label.config(text=f"{len(csv_files)} dosya dönüştürülüyor...")
            self.show_progress()
            
            # Threading kullan
            thread = threading.Thread(target=self.convert_folder, args=(folder_path, csv_files))
            thread.start()
    
    def convert_folder(self, folder_path, csv_files):
        """Klasör dönüştürme işlemi"""
        try:
            self.log(f"📂 Toplu işlem başladı: {len(csv_files)} dosya")
            
            table_style = self.get_selected_table_style()
            success_count = 0
            total_rows = 0
            
            for i, csv_file in enumerate(csv_files):
                csv_path = os.path.join(folder_path, csv_file)
                self.log(f"\n🔄 ({i+1}/{len(csv_files)}) {csv_file}")
                
                success, result, rows, cols = csv_to_formatted_xlsx(csv_path, None, table_style)
                
                if success:
                    self.log(f"✅ Başarılı: {rows} satır, {cols} sütun")
                    success_count += 1
                    total_rows += rows
                else:
                    self.log(f"❌ Hata: {result}")
            
            self.log(f"\n🎉 Toplu işlem tamamlandı!")
            self.log(f"✅ Başarılı: {success_count}/{len(csv_files)} dosya")
            self.log(f"📊 Toplam işlenen satır: {total_rows}")
            
            self.status_label.config(text="✅ Toplu dönüştürme tamamlandı!")
            messagebox.showinfo("Tamamlandı", f"Toplu işlem tamamlandı!\n\n{success_count}/{len(csv_files)} dosya başarıyla dönüştürüldü.")
            
        except Exception as e:
            self.log(f"❌ Toplu işlem hatası: {str(e)}")
            self.status_label.config(text="❌ Hata oluştu!")
            messagebox.showerror("Hata", f"Toplu işlem hatası:\n{str(e)}")
        
        finally:
            self.hide_progress()

def main():
    # Gerekli kütüphaneleri kontrol et
    try:
        import pandas
        import openpyxl
        import zipfile
    except ImportError as e:
        messagebox.showerror("Eksik Kütüphane", 
                           f"Gerekli kütüphane bulunamadı: {e}\n\n"
                           "Lütfen şu komutu çalıştırın:\n"
                           "pip install pandas openpyxl")
        return
    
    root = tk.Tk()
    app = CSVConverterGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()